"""Parsea un Excel de evento a la estructura JSON del sistema."""

from openpyxl import Workbook, load_workbook


def parse_excel(filepath):
    wb = load_workbook(filepath)

    # Hoja Evento
    ws = wb["Evento"]
    evento_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            evento_data[row[0].strip().lower()] = str(row[1]).strip()

    evento = {
        "nombre": evento_data.get("nombre", ""),
        "fecha": evento_data.get("fecha", ""),
        "lugar": evento_data.get("lugar", ""),
    }

    # Hoja Equipo
    ws = wb["Equipo"]
    headers = [c.value for c in ws[1]]
    base_cols = {"persona", "rol", "tarea", "hora", "depende de"}
    extra_cols = [h for h in headers if h and h.strip().lower() not in base_cols]

    personas = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        nombre = str(row[0]).strip()
        rol = str(row[1] or "").strip()
        tarea = str(row[2] or "").strip()
        hora = str(row[3] or "").strip()
        depende = str(row[4] or "").strip()

        if nombre not in personas:
            extras = {}
            for i, col_name in enumerate(extra_cols):
                idx = headers.index(col_name)
                val = row[idx] if idx < len(row) else None
                if val:
                    extras[col_name] = str(val).strip()
            personas[nombre] = {"nombre": nombre, "rol": rol, "tareas": [], "extras": extras}

        if tarea:
            personas[nombre]["tareas"].append({
                "nombre": tarea,
                "hora": hora,
                "estado": "pendiente",
                "timestamp": None,
                "depende_de": [d.strip() for d in depende.split(",") if d.strip()],
            })

    # Hoja Proveedores
    ws = wb["Proveedores"]
    prov_headers = [c.value for c in ws[1]]
    prov_base = {"servicio", "empresa"}
    prov_extra_cols = [h for h in prov_headers if h and h.strip().lower() not in prov_base]

    proveedores = []
    estado = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        servicio = str(row[0]).strip()
        empresa = str(row[1] or "").strip()
        extras = {}
        for col_name in prov_extra_cols:
            idx = prov_headers.index(col_name)
            val = row[idx] if idx < len(row) else None
            if val:
                extras[col_name] = str(val).strip()
        proveedores.append({"servicio": servicio, "empresa": empresa, "extras": extras})
        estado[servicio] = {"estado": "pendiente", "timestamp": None}

    return {
        "evento": evento,
        "equipo": list(personas.values()),
        "proveedores": proveedores,
        "estado": estado,
        "historial": [],
    }


def create_template(filepath):
    wb = Workbook()

    ws = wb.active
    ws.title = "Evento"
    ws.append(["Campo", "Valor"])
    ws.append(["Nombre", "Boda de X y Y"])
    ws.append(["Fecha", "12 Octubre 2026"])
    ws.append(["Lugar", "Hacienda Los Olivos"])

    ws = wb.create_sheet("Equipo")
    ws.append(["Persona", "Rol", "Tarea", "Hora", "Depende de"])
    ws.append(["Ana", "flores", "recepcion de flores", "09:00", ""])
    ws.append(["Ana", "flores", "montaje ceremonia", "11:00", "recepcion de flores"])

    ws = wb.create_sheet("Proveedores")
    ws.append(["Servicio", "Empresa"])
    ws.append(["flores", "Floreria Luna"])
    ws.append(["catering", "Banquetes del Valle"])

    wb.save(filepath)


def export_evento(data, filepath):
    wb = Workbook()

    # Evento
    ws = wb.active
    ws.title = "Evento"
    ws.append(["Campo", "Valor"])
    evento = data.get("evento", {})
    ws.append(["Nombre", evento.get("nombre", "")])
    ws.append(["Fecha", evento.get("fecha", "")])
    ws.append(["Lugar", evento.get("lugar", "")])

    # Equipo - collect all extra field names
    equipo = data.get("equipo", [])
    equipo_extra_keys = sorted({k for p in equipo for k in p.get("extras", {})})
    ws = wb.create_sheet("Equipo")
    ws.append(["Persona", "Rol", "Tarea", "Hora", "Estado", "Depende de"] + equipo_extra_keys)
    for p in equipo:
        extras = p.get("extras", {})
        extra_vals = [extras.get(k, "") for k in equipo_extra_keys]
        for t in p.get("tareas", []):
            ws.append([
                p["nombre"], p["rol"], t["nombre"], t["hora"],
                t.get("estado", "pendiente"),
                ", ".join(t.get("depende_de", [])),
            ] + extra_vals)
        if not p.get("tareas"):
            ws.append([p["nombre"], p["rol"], "", "", "", ""] + extra_vals)

    # Proveedores - collect all extra field names
    proveedores = data.get("proveedores", [])
    prov_extra_keys = sorted({k for pr in proveedores for k in pr.get("extras", {})})
    ws = wb.create_sheet("Proveedores")
    ws.append(["Servicio", "Empresa", "Estado"] + prov_extra_keys)
    estado = data.get("estado", {})
    for prov in proveedores:
        extras = prov.get("extras", {})
        extra_vals = [extras.get(k, "") for k in prov_extra_keys]
        s = estado.get(prov["servicio"], {})
        ws.append([prov["servicio"], prov["empresa"], s.get("estado", "")] + extra_vals)

    wb.save(filepath)
